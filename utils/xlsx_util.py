# -*- coding: UTF-8 -*-
'''
@Descripttion: deal with the xlsx file, convert to clean csv file
@version: 
@Author: Da Chuang
@Date: 2019-12-12 17:06:19
@LastEditors: Da Chuang
@LastEditTime: 2019-12-12 22:35:58
'''
import os
import sys
sys.path.insert(0, os.getcwd())
import pandas as pd
from openpyxl import load_workbook
import config

def get_clean_data():
    '''
    @Descripttion: 
    @param:
    @return: 
    '''
    wb = load_workbook(config.TWITTER_POLICES_NAME_FILE)
    # get the worksheet
    ws = wb.get_sheet_by_name(wb.sheetnames[0])
    clean_data = open(config.TWITTER_NAMES_FILE, 'w', encoding='utf-8-sig')
    for row in ws.rows:
        if(row[4].value == None):
            continue
        else:
            if row[1].value != None:
                country = row[1].value
            clean_data.writelines('%s,%s,%s,%s,%s\n' % (country,
                                                        str(row[2].value).replace(
                                                            ',', '.'),
                                                        str(row[3].value).replace(
                                                            ',', '.'),
                                                        str(row[4].value).replace(
                                                            ',', '.'),
                                                        str(row[5].value).replace(',', '.')))
    clean_data.close()


if __name__ == '__main__':
    # get_clean_data()
    pass
